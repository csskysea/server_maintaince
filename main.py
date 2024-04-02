
# -*- coding: utf-8 -*- 

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from mako.template import Template
from mako.lookup import TemplateLookup
import datetime
import shutil
import os.path
import openpyxl
import string,sys
import subprocess

class ServerManager(object):
    
    '''
    Get server information such as chassis NO. type,node NO.
    mac address,RAM, hostname from excel, and then update dhcp,
    slurm,dnsmasq configration, restart them to take effect.
    '''

    def __init__(self,excel_file):
        self.excel_file = excel_file
        self.server_info = []
        self.base_path = os.path.dirname(os.path.abspath( __file__ ))
        

    def getValueWithMergeLookup(self,mysheet, cell):
        idx = cell.coordinate
        merged_cells = list(mysheet.merged_cells)
        for row in merged_cells:        
            if idx in row:
                return mysheet.cell(row=row.min_row,column=row.min_col).value
            
        return mysheet[idx].value

    def get_server_info(self):
        
        wb = load_workbook(self.excel_file)
        node_sheet = wb['PhyNodes']
        list_ws_record = []
        #start from the second row and omit the first row, because the first row is title.
        for each_row in node_sheet.iter_rows(min_row=2):
            dict_row_record = {'chassis_num':'','type':'','node_num':'','cpu_count':'','core_count_percpu':'','mac':'','ip':'','dram_size':'','hostname':'','bmcip':''}
            # print([self.getValueWithMergeLookup(node_sheet,cell) for cell in each_row])
            for cell in each_row:
                if cell.column == 'A':
                    dict_row_record['chassis_num'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'B':
                    dict_row_record['type'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'C':
                    dict_row_record['node_num'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'D':
                    dict_row_record['cpu_count'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'E':
                    dict_row_record['core_count_percpu'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'F':
                    dict_row_record['mac'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'H':
                    dict_row_record['ip'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'I':
                    dict_row_record['dram_size'] = self.getValueWithMergeLookup(node_sheet,cell)
                elif cell.column == 'J':
                    dict_row_record['hostname'] = self.getValueWithMergeLookup(node_sheet,cell)            
                elif cell.column == 'K':
                    dict_row_record['bmcip'] = self.getValueWithMergeLookup(node_sheet,cell) 
            if dict_row_record['hostname'] == '' or dict_row_record['hostname'] == None:
                continue
            list_ws_record.append(dict_row_record)
        
        return list_ws_record
           
    def update_dhcp_cfg(self,list_server_info):
        '''
        the user who run the script must have root privilege.
        '''
        #back original configuration file first.(safety)

        lithium_dhcp_cfg = '/etc/dhcp/dhcpd.d/lithium.conf'
        lithium_dhcp_org_path = '/etc/dhcp/dhcpd.d/lithium_org'
        dhcpd_lithium_template_path = self.base_path
        if not os.path.exists(lithium_dhcp_org_path):
            os.mkdir(lithium_dhcp_org_path)
        if os.path.exists(lithium_dhcp_cfg):
            str_datatime = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            shutil.copyfile(lithium_dhcp_cfg,os.path.join(lithium_dhcp_org_path,str_datatime))
        lookup = TemplateLookup(
            directories=[dhcpd_lithium_template_path],
            output_encoding='utf-8',
            input_encoding='utf-8',
            default_filters=['decode.utf8'],
            encoding_errors='replace'
        )

        dhcpd_lithium_template = lookup.get_template('/dhcpd_lithium_template.conf')
        dhcpd_lithium_content = dhcpd_lithium_template.render(data=list_server_info)
        # print(dhcpd_lithium_content)
        with open(lithium_dhcp_cfg,'w+') as fd:
            fd.write(dhcpd_lithium_content)
            fd.close

    def update_slurm_cfg(self,list_server_info):
        '''
        the user who run the script must have root privilege.
        partition settings need you maintain 
        '''
        #back original configuration file first.(safety)
        slurm_conf = '/etc/slurm/slurm.conf'
        slurm_org_path = '/etc/slurm/slurm_org'
        slurm_template_path = self.base_path

        if not os.path.exists(slurm_org_path):
            os.mkdir(slurm_org_path)
        if os.path.exists(slurm_conf):
            str_datatime = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            shutil.copy(slurm_conf,os.path.join(slurm_org_path, str_datatime))

        lookup = TemplateLookup(
            directories=[slurm_template_path],
            output_encoding='utf-8',
            input_encoding='utf-8',
            default_filters=['decode.utf8'],
            encoding_errors='replace'
        )

        slurm_file_tempalte = lookup.get_template("/slurm_template.conf")                  
        slurm_file_content = slurm_file_tempalte.render(data=list_server_info)

        with open(slurm_conf,'w+') as fd:
            fd.write(slurm_file_content)
            fd.close
    
    def sync_hosts_file(self,lithium_conf_path):
        '''
        sync hosts file to lithium nodes.
        '''

        shutil.copy('/etc/hosts',os.path.join(lithium_conf_path,'hosts_conf'))
        
        
    def sync_slurm_cfg(self,lithium_conf_path):
        '''
        sync slurm.conf to lithium nodes.
        '''
        shutil.copy("/etc/slurm/slurm.conf",os.path.join(lithium_conf_path,'slurm_conf'))

    def update_hosts_file(self,list_server_info):
        hosts_path = '/etc/hosts'
        hosts_template_path = self.base_path
        lookup = TemplateLookup(
            directories=[hosts_template_path],
            output_encoding='utf-8',
            input_encoding='utf-8',
            default_filters=['decode.utf8'],
            encoding_errors='replace'
        )

        hosts_file_tempalte = lookup.get_template("/hosts_template.conf")                  
        hosts_file_content = hosts_file_tempalte.render(data=list_server_info)
        with open(hosts_path,'w+') as fd:
            fd.write(hosts_file_content)
            fd.close

    def update_maclist_file(self,list_server_info):
        maclist_path = "/etc/slurm/power/mac_list"
        maclist_template_path = self.base_path
        lookup = TemplateLookup(
            directories=[maclist_template_path],
            output_encoding='utf-8',
            input_encoding='utf-8',
            default_filters=['decode.utf8'],
            encoding_errors='replace'
        )
        maclist_file_tempalte = lookup.get_template("/maclist_template.conf")                  
        maclist_file_content = maclist_file_tempalte.render(data=list_server_info)
        with open(maclist_path,'w+') as fd:
            fd.write(maclist_file_content)
            fd.close

        
    def restart_dhcpd(self):
        '''
        the user who run the script must have root privilege.
        '''
        cmd_str = "service dhcpd restart"
        rtn_code = subprocess.call(cmd_str,shell=True)
        if(rtn_code != 0):
            print("restart dhcpd service failed,please check dhcp configuration!!")
            sys.exit()
            
    def restart_slurmctld(self):
        '''
        the user who run the script must have root privilege.
        you must restart lihtium nodes to use new slurm configuration.
        '''
        cmd_str = "service slrum restart"
        rtn_code = subprocess.call(cmd_str,shell=True)
        if(rtn_code != 0):
            print("restart slurmctld service failed,please check slurm configuration!!")
            sys.exit()

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("parameters count must be 2!!")
        sys.exit()
    rack_map_file = sys.argv[1]
    #lithium conf path:/home/public/lithium_conf/
    lithium_conf_path = sys.argv[2]
    server_mgr = ServerManager(rack_map_file)
    list_node_info = server_mgr.get_server_info()
    # print(list_node_info)
    print("Step1:update dhcp configuration for lithium nodes.")
    server_mgr.update_dhcp_cfg(list_node_info)
    print("restart dhcpd service.\n")
    server_mgr.restart_dhcpd()
    print("\nStep2:update hosts file for lithium nodes.\n")
    server_mgr.update_hosts_file(list_node_info)
    print("...sync to lithium nodes.\n")
    server_mgr.sync_hosts_file(lithium_conf_path)
    print("\nStep3:update slurm configuration for lithium nodes.\n")
    server_mgr.update_slurm_cfg(list_node_info)
    print("...sync to lithium nodes.\n")
    server_mgr.sync_slurm_cfg(lithium_conf_path)
    print("...restart slurmctld daemon.\n")
    server_mgr.restart_slurmctld()
    server_mgr.update_maclist_file(list_node_info)
    
